import streamlit as st
import pandas as pd
import numpy as np
import io
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------- PAGE CONFIG ----------------
st.set_page_config(page_title="Coffee Trade Intelligence", layout="wide")

# ---------------- GLOBAL CSS (UNCHANGED UI) ----------------
st.markdown("""
<style>
.stApp {
    background: linear-gradient(180deg, #020617 0%, #020617 100%);
    color: white;
    font-family: 'Inter', sans-serif;
}
.hero {
    background-image: linear-gradient(rgba(2,6,23,0.75), rgba(2,6,23,0.95)),
        url("https://images.unsplash.com/photo-1498804103079-a6351b050096");
    background-size: cover;
    background-position: center;
    padding: 80px 40px;
    border-radius: 16px;
    margin-bottom: 40px;
}
.hero h1 { font-size: 42px; font-weight: 600; }
.hero p { color: #94a3b8; }
.pipeline { padding: 20px; }
.block {
    background: #0f172a;
    padding: 25px;
    border-radius: 12px;
    margin-bottom: 20px;
}
.stButton button {
    background: #1d4ed8;
    color: white;
    border-radius: 8px;
}
</style>
""", unsafe_allow_html=True)

# ---------------- HERO ----------------
st.markdown("""
<div class="hero">
    <h1>Coffee Trade Intelligence</h1>
    <p>Upload → Clean → Convert → Download</p>
</div>
""", unsafe_allow_html=True)

# ================================================================
# CONSTANTS — HSN BUCKETS
# ================================================================

# Target soluble coffee codes
SOLUBLE_COFFEE_HSN = {21011110, 21011120, 21011130, 21011190}

# Chicory/coffee premix blends
CHICORY_PREMIX_HSN = {21011200}

# All primary target codes
ALL_TARGET_HSN = SOLUBLE_COFFEE_HSN | CHICORY_PREMIX_HSN

# Standalone chicory-only exports — get their own sheet, never mixed with soluble coffee
CHICORY_ONLY_HSN = {21013010}

# Pure chicory extracts / green coffee extract — excluded from soluble coffee,
# but DO NOT scan these for wrong-HSN soluble coffee either
PURE_CHICORY_HSN = {21013090, 21012020}

# Ground/roasted coffee — NEVER soluble, never scan
GROUND_COFFEE_HSN = {9012190, 9019090, 9019020, 9011119, 9011129, 9012290}

# HSN codes that should NEVER be scanned for wrong-HSN soluble coffee
# (pure chicory extracts + ground roasted coffee)
EXCLUDE_SCAN_HSN = PURE_CHICORY_HSN | GROUND_COFFEE_HSN

# HSN codes to rescue from wrong-HSN scanner — confirmed systemic misfiling
# These exporters habitually file soluble coffee under roasted/ground codes
RESCUE_HSN = {9012190, 9019090, 9019020, 9011119, 9011129, 9012290}
# Note: RESCUE_HSN is a subset of GROUND_COFFEE_HSN — we selectively scan these
# because the audit showed consistent soluble coffee misfiling across 25 files.
# They are removed from EXCLUDE_SCAN_HSN below.
EXCLUDE_SCAN_HSN = PURE_CHICORY_HSN  # Only pure chicory codes are truly never-scan

# ================================================================
# CONSTANTS — KEYWORD PATTERNS
# ================================================================

# Soluble coffee detection — COFFEE-ANCHORED to avoid false positives
# from freeze-dried herbs, ginger, dill, coriander etc.
# Key fix: "FREEZE DRIED" and "SPRAY DRIED" alone are too broad.
# Must be paired with COFFEE to qualify.
SOLUBLE_KEYWORDS = [
    'INSTANT COFFEE',
    'SOLUBLE COFFEE',
    'SPRAY DRIED COFFEE',
    'FREEZE DRIED COFFEE',
    'AGGLOMERATED COFFEE',
    'AGGLOMERATED INSTANT',
    'FREEZE-DRIED COFFEE',
    'SPRAY-DRIED COFFEE',
    'COFFEE EXTRACT POWDER',
    'COFFEE PREMIX',
    'NESCAFE',
    'BRU INSTANT',
    'SUNRISE EXTRA',
]

# Chicory detection for wrong-HSN scanner
CHICORY_WRONG_HSN_KEYWORDS = ['CHICORY', 'CHICCORY']

sol_pattern = '|'.join(re.escape(k) for k in SOLUBLE_KEYWORDS)
chic_wrong_pat = '|'.join(re.escape(k) for k in CHICORY_WRONG_HSN_KEYWORDS)

# ================================================================
# CONSTANTS — EXCLUSION KEYWORDS (HARDCODED)
# These are patterns confirmed across 25 files to always be non-soluble.
# They are hardcoded because they are structural, not exporter-specific.
# ================================================================
HARDCODED_EXCLUSIONS = [
    # Non-soluble coffee forms
    ('GINGER COFFEE', ''),
    ('CHUKKU', ''),
    ('SUKKU', ''),
    ('KAPI', ''),
    ('KAPPI', ''),
    ('KAAPI', ''),
    ('FILTER COFFEE POWDER', ''),
    ('GROUND COFFEE', ''),
    ('ROASTED COFFEE', ''),
    # Cold coffee / liquid — not soluble powder
    ('COLD COFFEE PREMIX', ''),       # systemic false positive across 9011129, 9019020, 9012190
    ('COLD COFFEE', ''),
    ('LIQUID COFFEE', ''),
    ('DECOCTION', ''),
    ('COLD BREWED COFFEE CONCENTRATE', ''),
    # Flavoured drinks
    ('FRAPPE', ''),
    ('CHOCOLATE MILKSHAKE', ''),
    ('MOJITO', ''),
    ('MASALA LEMONADE', ''),
    ('HOT CHOCOLATE', ''),
    ('COLD COCO', ''),
    # Non-food / pharma / industrial
    ('CAPSICUM POWDER', ''),
    ('FREEZE DRIED SHRIMP', ''),
    ('FREEZE DRIED GREEN PEPPER', ''),
    ('FREEZE DRIED RED PEPPER', ''),
    ('FREEZE DRIED DILL', ''),
    ('FREEZE DRIED CORIANDER', ''),
    ('FREEZE DRIED THYME', ''),
    ('FREEZE DRIED GINGER', ''),
    ('FREEZE DRIED CURRY LEAVES', ''),
    ('FREEZE DRIED CILANTRO', ''),
    ('FREEZE DRIED JAMUN', ''),
    ('FREEZE DRIED RED BELL PEPPER', ''),
    ('SPRAY DRIED GINGER', ''),
    ('SPRAY DRIED SOY SAUCE', ''),
    ('SPRAY DRIED H.V.P', ''),
    ('SPRAY DRIED HVP', ''),
    ('SPRAY DRIED BLACK TEA', ''),
    ('PAPAYA POWDER', ''),
    ('MANGO POWDER', ''),
    ('CYPROHEPTADINE', ''),
    ('GINKGO BILOBA', ''),
    ('CARBONYL IRON', ''),
    ('NON DAIRY CREAMER', ''),
    ('H.V.P POWDER', ''),
    ('HVP POWDER', ''),
    ('SUGAR FREE DLITE', ''),
    ('CHLOROGENIC ACID', ''),
    ('NATURAL ANHYDROUS CAFFEINE', ''),
    ('CAFFEINE ANHYDROUS', ''),
    ('NATURAL CAFFEINE', ''),
    ('VEGETABLE FAT', ''),
    ('PALM FAT', ''),
    # Green coffee / herbal extracts (nutraceutical, not soluble)
    ('GREEN COFFEE EXTRACT', ''),
    ('GREEN COFFEE BEAN EXTRACT', ''),
    ('GREEN COFFEE ROBUSTA', ''),
    ('COFFEE BEAN EXTRACT', ''),
    ('COFFEE BERRY EXTRACT', ''),
    ('COFFEE OIL', ''),
    # Tea and tea premixes — global
    ('INSTANT TEA', ''),
    ('MASALA TEA', ''),
    ('MASALATEA', ''),          # SENSO exporter no-space variant
    ('MASALA CHAI', ''),
    ('CARDAMON TEA', ''),
    ('CARDAMOM CHAI', ''),
    ('CARDAMOM PREMIX TEA', ''),
    ('CHAI MIX', ''),
    ('HERBAL TEA', ''),
    ('LEMON ICED TEA', ''),
    ('PEACH ICED TEA', ''),
    ('GREEN TEA EXTRACT', ''),
    ('PURPLE TEA EXTRACT', ''),
    ('JAVA TEA EXTRACT', ''),
    ('TURM', ''),               # TURMERIC-based tea/latte
    ('TULSI', ''),
    ('MINT GREEN TEA', ''),
    # HSN-specific: 21011190 herbal extracts
    ('HERBAL EXTRACT', '21011190'),
    ('HERBALEXTRACTS', '21011190'),   # no-space exporter variant
    ('GUAYUSA', '21011190'),
    ('BUDAN HEAD', '21011190'),
    # Tea brands
    ('KESAR BADAM', ''),
    ('ADRAK', ''),
    ('3 ROSES', '21011110'),
    ('RED LABEL', '21011110'),
    ('TAJ MAHAL', '21011110'),
    ('SWAN', '21011120'),
    ('WAGH BAKRI', ''),
    ('WAGHBAKRI', ''),
    ('LIPTON', ''),
    ('NESTEA', ''),
    ('BROOKE BOND', ''),
    ('KANNAN DEVAN', ''),
    # Herbal / ayurvedic
    ('HERBAL PRODUCTS', ''),
    ('SORIG', ''),
    ('HERBAL COFFEE', ''),
    # Pods
    ('DOLCE GUSTO', ''),
    ('NESPRESSO PODS', ''),
    # Other drinks
    ('MATCHA', ''),
    ('SODA', ''),
    ('MANGO MATCHA', ''),
    ('TURERIC LATTE', ''),
    ('TURMERIC LATTE', ''),
    # Noodles
    ('MASALA NOODLES', ''),
    ('INSTANT NOODLES', ''),
    ('NOODLES', '21011200'),
    ('ATTA NOODLES', ''),
    # Misc non-product rows
    ('TAX INVOICE', ''),
    ('FREE SAMPLE', ''),
    # Industrial
    ('DRIVE 68', ''),
    # 21011200 specific: lemongrass tea, dry ginger cappi
    ('LEMON GRASS TEA', '21011200'),
    ('DRY GINGER CAPPI', '21011200'),
    # 21012010 specific: tea extracts misfiled
    ('SPRAY DRIED BLACK TEA EXTRACT', ''),
    ('LEMONICED TEA', ''),
    ('PEACHICED TEA', ''),
    # Ketchup / food bundling
    ('KETCHUP', ''),
    ('HOT & SWEET', ''),
    ('TOMATO', ''),
]

# ================================================================
# PRE-COMPILED PATTERNS — built once at startup, reused every file
# ================================================================

# Global hardcoded exclusion pattern (no HSN filter)
_global_excl_pattern = re.compile(
    '|'.join(re.escape(k) for k, h in HARDCODED_EXCLUSIONS if h == ''),
    re.IGNORECASE
)

# Per-HSN hardcoded exclusion patterns — dict keyed by HSN string
_hsn_excl_patterns = {}
for _kw, _hsn in HARDCODED_EXCLUSIONS:
    if _hsn:
        if _hsn not in _hsn_excl_patterns:
            _hsn_excl_patterns[_hsn] = []
        _hsn_excl_patterns[_hsn].append(re.escape(_kw))
_hsn_excl_compiled = {
    hsn: re.compile('|'.join(patterns), re.IGNORECASE)
    for hsn, patterns in _hsn_excl_patterns.items()
}

# Strict coffee signal pattern for HSN 21011190 / 21011200
_strict_coffee_signal_pattern = re.compile(
    r'COFFEE|CAPPUCCINO|CAPUCCINO|NESCAFE|BRU|LEVISTA|COTHAS|CONTINENTAL|NARASUS|TATA COFFEE|CHICORY|PREMIX',
    re.IGNORECASE
)
STRICT_CHECK_HSNS = {'21011190', '21011200'}


def build_excl_list_lookup(excl_df):
    """
    Pre-process the user exclusion list into two structures for O(1) lookup:
      - global_keywords: set of (keyword,) tuples with no HSN filter
      - hsn_keywords: dict of hsn_str -> list of (keyword, reason)
    Returns (global_set, hsn_dict, global_reasons_dict)
    """
    global_kws = {}   # keyword -> reason
    hsn_kws    = {}   # hsn_str -> {keyword: reason}

    for _, r in excl_df.iterrows():
        kw  = str(r.get('KEYWORD', '')).upper().strip()
        hsn = str(r.get('HSN_FILTER', '')).strip()
        reason = str(r.get('REASON', 'Exclusion list match'))
        if not kw:
            continue
        if hsn == '' or hsn == 'nan' or hsn == 'NAN':
            global_kws[kw] = reason
        else:
            hsn_kws.setdefault(hsn, {})[kw] = reason

    return global_kws, hsn_kws


# ================================================================
# VECTORISED EXCLUSION — operates on entire DataFrame columns at once
# No Python loops over rows. Called once per file.
# ================================================================
def apply_exclusions_vectorised(df, excl_global_kws, excl_hsn_kws):
    """
    Returns two Series: _EXCLUDED (bool) and _EXCL_REASON (str).
    All operations are pandas/numpy vectorised — no row-level Python loops.
    """
    desc  = df['_DESC_UP']           # already upper-stripped
    hsn_s = df['_HSN_INT'].astype(str)

    n = len(df)
    excluded = np.zeros(n, dtype=bool)
    reason   = np.full(n, '', dtype=object)

    # ── 1. Hardcoded global pattern (single compiled regex on whole column) ──
    hit = desc.str.contains(_global_excl_pattern, na=False)
    excluded |= hit.values
    reason[hit.values] = 'Hardcoded exclusion'

    # ── 2. Hardcoded HSN-specific patterns ──
    for hsn_str, pat in _hsn_excl_compiled.items():
        mask = (hsn_s == hsn_str) & (~excluded)
        if not mask.any():
            continue
        hit = desc[mask].str.contains(pat, na=False)
        idx = hit[hit].index
        excluded[df.index.get_indexer(idx)] = True
        reason[df.index.get_indexer(idx)]   = 'Hardcoded exclusion (HSN-specific)'

    # ── 3. User exclusion list — global keywords ──
    if excl_global_kws:
        # Build a single regex from all global user keywords
        user_global_pat = re.compile(
            '|'.join(re.escape(k) for k in excl_global_kws),
            re.IGNORECASE
        )
        mask = ~excluded
        if mask.any():
            hit = desc[mask].str.contains(user_global_pat, na=False)
            # For reason: find first matching keyword per row (cheap since list is small)
            for kw, rsn in excl_global_kws.items():
                kw_hit = desc[mask].str.contains(re.escape(kw), na=False, regex=True)
                idx = kw_hit[kw_hit & ~pd.Series(excluded[df.index.get_indexer(mask[mask].index)],
                             index=mask[mask].index)].index
                pos = df.index.get_indexer(idx)
                excluded[pos] = True
                for p in pos:
                    if reason[p] == '':
                        reason[p] = rsn

    # ── 4. User exclusion list — HSN-specific keywords ──
    for hsn_str, kw_dict in excl_hsn_kws.items():
        hsn_mask = (hsn_s == hsn_str) & (~pd.Series(excluded, index=df.index))
        if not hsn_mask.any():
            continue
        for kw, rsn in kw_dict.items():
            hit = desc[hsn_mask].str.contains(re.escape(kw), na=False, regex=True)
            idx = hit[hit].index
            pos = df.index.get_indexer(idx)
            excluded[pos] = True
            for p in pos:
                if reason[p] == '':
                    reason[p] = rsn

    # ── 5. Strict coffee signal check for 21011190 / 21011200 ──
    strict_mask = hsn_s.isin(STRICT_CHECK_HSNS) & (~pd.Series(excluded, index=df.index))
    if strict_mask.any():
        has_signal = desc[strict_mask].str.contains(_strict_coffee_signal_pattern, na=False)
        no_signal  = strict_mask & (~has_signal.reindex(df.index, fill_value=False))
        pos = df.index.get_indexer(no_signal[no_signal].index)
        excluded[pos] = True
        for p in pos:
            if reason[p] == '':
                reason[p] = 'No coffee signal — strict check for HSN'

    return pd.Series(excluded, index=df.index), pd.Series(reason, index=df.index)

# ================================================================
# HSN BUCKETING — vectorised
# ================================================================
def norm_hsn_series(series):
    """Vectorised HSN normalisation — no Python loop."""
    return (
        series.astype(str)
              .str.replace(r'\s+', '', regex=True)
              .str.split('.').str[0]
              .pipe(pd.to_numeric, errors='coerce')
              .fillna(0)
              .astype(int)
    )

def bucket_hsn_series(hsn_series):
    """Vectorised HSN bucketing using np.select — much faster than .apply()."""
    conditions = [
        hsn_series.isin(SOLUBLE_COFFEE_HSN),
        hsn_series.isin(CHICORY_PREMIX_HSN),
        hsn_series.isin(CHICORY_ONLY_HSN),
        hsn_series.isin(PURE_CHICORY_HSN),
        hsn_series.isin(GROUND_COFFEE_HSN),
    ]
    choices = ['SOLUBLE_COFFEE', 'CHICORY_PREMIX', 'CHICORY_ONLY', 'PURE_CHICORY', 'GROUND_COFFEE']
    return pd.Series(np.select(conditions, choices, default='OTHER'), index=hsn_series.index)

# Keep scalar versions for any legacy calls
def norm_hsn(v):
    try:
        return int(str(v).replace(' ', '').strip().split('.')[0])
    except:
        return 0

# ================================================================
# WRONG HSN SCANNER
# Key fixes:
#   - 21013090 (pure chicory extracts) is in EXCLUDE_SCAN_HSN — never scanned
#   - GROUND_COFFEE_HSN codes ARE scanned (audit showed systemic misfiling)
#   - sol_pattern is now coffee-anchored (FREEZE DRIED COFFEE, not FREEZE DRIED)
# ================================================================
def find_wrong_hsn_rows(df, hs_col):
    # Rows not in primary target AND not pure chicory (21013090, 21012020)
    scannable_mask = (
        ~df['_BUCKET'].isin(['SOLUBLE_COFFEE', 'CHICORY_PREMIX', 'CHICORY_ONLY', 'PURE_CHICORY'])
    )
    df_scan = df[scannable_mask].copy()

    df_scan['_WRONG_SOLUBLE'] = df_scan['_DESC_UP'].str.contains(
        sol_pattern, na=False, regex=True
    )
    df_scan['_WRONG_CHICORY'] = df_scan['_DESC_UP'].str.contains(
        chic_wrong_pat, na=False, regex=True
    )
    df_scan['_WRONG_ANY'] = df_scan['_WRONG_SOLUBLE'] | df_scan['_WRONG_CHICORY']

    return df_scan[df_scan['_WRONG_ANY']].copy()

# ================================================================
# KNOWN BRAND CHICORY CLASSIFICATION TABLE
# ================================================================
KNOWN_BRANDS = [
    (r'NESCAFE.*SUNRISE|SUNRISE.*REGULAR|SUNRISE EXTRA|SUNRISE BLENDED|SUNRISE INSTA', 70, 30, 'CONFIRMED', 'Nestle Professional listing'),
    (r'NESCAFE CLASSIC', 100, 0, 'CONFIRMED', 'Pure instant coffee'),
    (r'NESCAFE GOLD', 100, 0, 'CONFIRMED', 'Premium pure coffee'),
    (r'NESCAFE INTENSO', 100, 0, 'CONFIRMED', 'Pure instant — confirmed'),
    (r'NESCAFE', 100, 0, 'ASSUMED', 'Generic Nescafe — assumed pure unless variant known'),
    (r'BRU.*GOLD|BRU GOLD', 100, 0, 'CONFIRMED', 'Pure freeze dried'),
    (r'BRU.*GREEN LABEL|GREEN LABEL.*COFFEE', 80, 20, 'ASSUMED', 'Filter blend'),
    (r'BRU.*SELECT', 85, 15, 'ASSUMED', 'Premium blend'),
    (r'BRU INSTANT|BRU.*OPTIMA|BRU.*OPTM|BRU.*SUPER STRONG|BRU STRONG|BRU.*PLATINA|BRU.*EXPORT|BRU.*STAND UP|BRU.*AROMA|BRU.*INST', 70, 30, 'ASSUMED', 'Known chicory blend'),
    (r'TATA.*GRAND|TATA COFFEE GRAND', 70, 30, 'ASSUMED', 'Chicory mix indicated'),
    (r'TATA.*GOLD|TATA COFFEE GOLD', 100, 0, 'CONFIRMED', 'Pure coffee'),
    (r'TATA.*CLASSIC', 100, 0, 'ASSUMED', 'Usually pure'),
    (r'CONTINENTAL.*MALGUDI', 53, 47, 'CONFIRMED', 'Label states 53:47'),
    (r'CONTINENTAL.*XTRA|CONTINENTAL.*STRONG', 70, 30, 'ASSUMED', 'Tradeindia listing'),
    (r'CONTINENTAL.*SPECIAL|CONTINENTAL.*PURE', 100, 0, 'CONFIRMED', 'Pure variant'),
    (r'LEVISTA.*CLASSIC', 80, 20, 'ASSUMED', 'Chicory variant'),
    (r'LEVISTA.*80', 80, 20, 'ASSUMED', 'Blend ratio in name'),
    (r'LEVISTA.*70', 70, 30, 'ASSUMED', 'Blend ratio in name'),
    (r'LEVISTA.*60', 60, 40, 'ASSUMED', 'Blend ratio in name'),
    (r'LEVISTA.*PREMIUM|LEVISTA.*PURE', 100, 0, 'ASSUMED', 'Pure line'),
    (r'NARASUS.*UDHAYAM|NARASUS.*UDHAIYAM', 80, 20, 'ASSUMED', 'Blend positioning'),
    (r'NARASUS.*DELITE', 55, 45, 'ASSUMED', 'Label: 55:45'),
    (r'NARASUS.*BESH SUKKU|BESH SUKKU', 70, 30, 'ASSUMED', 'Sukku blend'),
    (r'NARASUS.*PURE|NARASUS PURE INSTANT|NARASUS INSTA STRONG|NARASUS STRONG INSTANT|NARASUS INSTANT', 100, 0, 'ASSUMED', 'Pure/instant line'),
    (r'SUNRISE COFFEE|SUNRISE.*BLENDED', 70, 30, 'CONFIRMED', 'Nestle Professional listing'),
    (r'COTHAS.*SPECIAL', 80, 20, 'ASSUMED', 'Special filter blend'),
    (r'COTHAS.*PREMIUM', 85, 15, 'ASSUMED', 'Premium blend'),
    (r'COTHAS.*80', 80, 20, 'ASSUMED', 'Ratio in name'),
    (r'KDC.*60|KDC.*70|KDC.*80', None, None, 'EXPLICIT_RATIO', 'Ratio in product name — use extract_ratio'),
]

def extract_ratio(desc_up):
    m = re.search(r'\b(\d{2})\s*[:/]\s*(\d{2})\b', desc_up)
    if m:
        return int(m.group(1)), int(m.group(2))
    return None, None

def classify_chicory_row(desc_up):
    has_chicory_word = bool(re.search(r'CHICORY|CHICCORY|CICCORY|RICORY', desc_up))
    ratio_a, ratio_b = extract_ratio(desc_up)

    # EXPLICIT: stated in description
    if has_chicory_word and ratio_a is not None:
        return ('EXPLICIT', ratio_a, ratio_b, 'HIGH', 'Chicory + ratio both stated in description')
    if has_chicory_word:
        return ('EXPLICIT', None, None, 'MEDIUM', 'Chicory stated, ratio not given')
    if ratio_a is not None:
        return ('EXPLICIT', ratio_a, ratio_b, 'MEDIUM', 'Ratio in description, chicory not named')

    # KNOWN BRAND
    for pattern, c_pct, ch_pct, conf, notes in KNOWN_BRANDS:
        if re.search(pattern, desc_up):
            if conf == 'EXPLICIT_RATIO':
                r_a, r_b = extract_ratio(desc_up)
                if r_a:
                    return ('EXPLICIT', r_a, r_b, 'HIGH', notes)
            if ch_pct == 0:
                return ('PURE_COFFEE', c_pct, ch_pct, conf, notes)
            return ('KNOWN_BRAND', c_pct, ch_pct, conf, notes)

    return None

# ================================================================
# CHICORY SIGNAL PATTERN
# ================================================================
CHICORY_SIGNAL_PATTERN = (
    r'CHICORY|CHICCORY|CICCORY|RICORY|'
    r'\b\d{2}\s*[:/]\s*\d{2}\b|'
    r'SUNRISE EXTRA|SUNRISE.*BLENDED|SUNRISE.*INSTA|'
    r'BRU INSTANT|BRU.*OPTIMA|BRU.*OPTM|BRU.*SUPER STRONG|BRU STRONG|BRU.*PLATINA|BRU.*AROMA|'
    r'BRU.*INST(?!.*GOLD)|'
    r'TATA.*GRAND|CONTINENTAL.*MALGUDI|CONTINENTAL.*XTRA|CONTINENTAL.*STRONG|'
    r'NARASUS.*UDHAYAM|NARASUS.*DELITE|NARASUS.*BESH SUKKU|'
    r'LEVISTA.*CLASSIC|LEVISTA.*[678]0'
)

# ================================================================
# MT CONVERSION — vectorised (logic unchanged, no apply loop)
# ================================================================
def convert_to_mt(row):
    """Scalar version kept for reference — not called in hot path."""
    qty = row.get('STANDARD QUANTITY')
    unit = str(row.get('STANDARD QUANTITY UNIT', '')).upper()
    if pd.isna(qty):
        return None, 'BLANK'
    try:
        qty = float(qty)
    except:
        return None, 'BLANK'
    if unit in ('KGS', 'KG'):
        return qty / 1000, 'DIRECT'
    if unit in ('MTS', 'MT'):
        return qty, 'DIRECT'
    return None, 'BLANK'

def convert_to_mt_vectorised(df):
    """
    Vectorised MT conversion. Returns (mt_series, status_series).
    Logic is identical to convert_to_mt — just no Python loop.
    """
    qty  = pd.to_numeric(df.get('STANDARD QUANTITY', pd.Series(dtype=float)), errors='coerce')
    unit = df.get('STANDARD QUANTITY UNIT', pd.Series(dtype=str)).astype(str).str.upper().str.strip()

    is_blank  = qty.isna()
    is_kgs    = unit.isin(['KGS', 'KG']) & ~is_blank
    is_mt     = unit.isin(['MTS', 'MT']) & ~is_blank

    mt_vals = np.where(is_kgs, qty / 1000,
              np.where(is_mt,  qty,
                       np.nan))

    status = np.where(is_blank, 'BLANK',
             np.where(is_kgs | is_mt, 'DIRECT',
                      'BLANK'))

    return (
        pd.Series(np.where(np.isnan(mt_vals), None, mt_vals), index=df.index),
        pd.Series(status, index=df.index)
    )

# ================================================================
# EXCEL FORMATTING (same colour palette as original)
# ================================================================
SHEET_COLOURS = {
    '1 All Soluble Coffee':      ('1F4E79', 'D6E4F0', 'EBF4FA', 'FFFFFF'),
    '2 Chicory Explicit Ratio':  ('1A5E20', 'D4EDDA', 'EAF7EC', 'FFFFFF'),
    '3 Chicory Known Brand':     ('7B4F00', 'FFF3CD', 'FFFAED', 'FFFFFF'),
    '4 Chicory Assumed':         ('6A0572', 'F3D6F5', 'FAF0FB', 'FFFFFF'),
    '5 Chicory Only Exports':    ('2C3E50', 'D5DBDB', 'F2F3F4', 'FFFFFF'),
    '6 Excluded Items Review':   ('4A4A4A', 'E8E8E8', 'F5F5F5', 'FFFFFF'),
    'Summary':                   ('0D3B66', 'D6E4F0', 'EBF4FA', 'FFFFFF'),
}

def hex_fill(hex_code):
    return PatternFill(start_color=hex_code, end_color=hex_code, fill_type='solid')

def thin_border():
    side = Side(style='thin', color='CCCCCC')
    return Border(left=side, right=side, top=side, bottom=side)

def format_sheet(ws, header_hex, row1_hex, row2_hex, font_hex):
    header_fill = hex_fill(header_hex)
    header_font = Font(bold=True, color=font_hex, size=10)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border()

    fill1 = hex_fill(row1_hex)
    fill2 = hex_fill(row2_hex)
    data_font = Font(size=9, color='1A1A1A')
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=1):
        fill = fill1 if row_idx % 2 == 1 else fill2
        for cell in row:
            cell.fill = fill
            cell.font = data_font
            cell.alignment = Alignment(vertical='center', wrap_text=False)
            cell.border = thin_border()

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    ws.freeze_panes = 'A2'

# ================================================================
# MAIN PROCESSING FUNCTION
# ================================================================
DROP_COLS = ['_HSN_INT', '_DESC_UP', '_BUCKET', '_EXCLUDED', '_EXCL_REASON',
             '_IS_CHICORY_SIGNAL', '_WRONG_SOLUBLE', '_WRONG_CHICORY', '_WRONG_ANY',
             '_SOURCE_TEMP']

def clean_export(df_in):
    keep = [c for c in df_in.columns if c not in DROP_COLS]
    return df_in[keep].reset_index(drop=True)

def process_file(file, excl_df):
    # ── READ — explicit engine is faster than auto-detect ──
    df = pd.read_excel(file, engine='openpyxl')
    df.columns = df.columns.str.strip()

    # ── DETECT COLUMNS ──
    hs_col = next((c for c in df.columns if 'HS' in c.upper() and 'CODE' in c.upper()), None)
    if not hs_col:
        hs_col = next((c for c in df.columns if 'HS' in c.upper()), None)
    desc_col = next((c for c in df.columns if 'PRODUCT' in c.upper() and 'DESC' in c.upper()), None)
    if not desc_col:
        desc_col = next((c for c in df.columns if 'DESC' in c.upper()), None)

    if not hs_col or not desc_col:
        st.error(f"Could not find HS CODE or PRODUCT DESCRIPTION columns in {file.name}")
        return None

    # ── NORMALISE — vectorised, no .apply() ──
    df['_HSN_INT'] = norm_hsn_series(df[hs_col])
    df['_DESC_UP'] = df[desc_col].astype(str).str.upper().str.strip()
    df['_BUCKET']  = bucket_hsn_series(df['_HSN_INT'])

    # ── EXCLUSIONS — fully vectorised, build lookup once per file ──
    excl_global_kws, excl_hsn_kws = build_excl_list_lookup(excl_df)
    df['_EXCLUDED'], df['_EXCL_REASON'] = apply_exclusions_vectorised(
        df, excl_global_kws, excl_hsn_kws
    )

    # ── WRONG HSN SCANNER — already vectorised via str.contains ──
    df_wrong = find_wrong_hsn_rows(df, hs_col)
    df_wrong = df_wrong[~df_wrong['_EXCLUDED']].copy()
    df_wrong['_SOURCE'] = 'Wrong HSN — flagged'

    # ── SHEET 1: ALL SOLUBLE COFFEE ──
    df_correct = df[
        df['_BUCKET'].isin(['SOLUBLE_COFFEE', 'CHICORY_PREMIX']) & ~df['_EXCLUDED']
    ].copy()
    df_correct['_SOURCE'] = 'Correct HSN'

    df_sheet1 = pd.concat([df_correct, df_wrong], ignore_index=True)

    # ── MT CONVERSION — vectorised, no .apply() ──
    df_sheet1['MT'], df_sheet1['MT_STATUS'] = convert_to_mt_vectorised(df_sheet1)

    # ── CHICORY SIGNAL DETECTION — vectorised str.contains ──
    df_sheet1['_IS_CHICORY_SIGNAL'] = df_sheet1['_DESC_UP'].str.contains(
        CHICORY_SIGNAL_PATTERN, na=False, regex=True
    )

    df_chicory_pool = df_sheet1[df_sheet1['_IS_CHICORY_SIGNAL']].copy()

    def apply_classification(df_in):
        # classify_chicory_row uses .apply() but only on ~200-400 chicory rows,
        # not the full 70k — acceptable cost
        results = df_in['_DESC_UP'].apply(classify_chicory_row)
        df_in = df_in.copy()
        df_in['BLEND_CATEGORY'] = results.apply(lambda x: x[0] if x else 'ASSUMED')
        df_in['COFFEE_PCT']     = results.apply(lambda x: x[1] if x else None)
        df_in['CHICORY_PCT']    = results.apply(lambda x: x[2] if x else None)
        df_in['CONFIDENCE']     = results.apply(lambda x: x[3] if x else 'LOW')
        df_in['BLEND_NOTES']    = results.apply(lambda x: x[4] if x else 'Chicory signal but no confirmed ratio')
        return df_in

    df_chicory_pool = apply_classification(df_chicory_pool)

    df_premix = df_sheet1[
        (df_sheet1['_BUCKET'] == 'CHICORY_PREMIX') &
        (~df_sheet1.index.isin(df_chicory_pool.index))
    ].copy()
    df_premix = apply_classification(df_premix)

    df_all_chicory = pd.concat([df_chicory_pool, df_premix], ignore_index=True)

    df_chic_explicit = df_all_chicory[df_all_chicory['BLEND_CATEGORY'] == 'EXPLICIT'].copy()
    df_chic_known    = df_all_chicory[df_all_chicory['BLEND_CATEGORY'] == 'KNOWN_BRAND'].copy()
    df_chic_assumed  = df_all_chicory[df_all_chicory['BLEND_CATEGORY'].isin(['ASSUMED', 'PURE_COFFEE'])].copy()

    # ── SHEET 5: CHICORY ONLY EXPORTS (21013010) ──
    df_chicory_only = df[df['_BUCKET'] == 'CHICORY_ONLY'].copy()
    if len(df_chicory_only):
        df_chicory_only['MT'], df_chicory_only['MT_STATUS'] = convert_to_mt_vectorised(df_chicory_only)

    # ── SHEET 6: EXCLUDED ITEMS REVIEW ──
    df_excl_from_target = df[
        df['_BUCKET'].isin(['SOLUBLE_COFFEE', 'CHICORY_PREMIX']) & df['_EXCLUDED']
    ].copy()
    df_excl_from_target['_EXCL_SOURCE'] = 'Keyword exclusion — target HSN'

    scannable_mask = ~df['_BUCKET'].isin([
        'SOLUBLE_COFFEE', 'CHICORY_PREMIX', 'CHICORY_ONLY', 'PURE_CHICORY'
    ])
    df_other_not_flagged = df[scannable_mask].copy()
    df_other_not_flagged = df_other_not_flagged[
        ~df_other_not_flagged.index.isin(df_wrong.index)
    ].copy()
    df_other_not_flagged['_EXCL_SOURCE'] = 'Other HSN — not flagged as soluble'

    df_excluded_review = pd.concat(
        [df_excl_from_target, df_other_not_flagged], ignore_index=True
    )

    # ── SUMMARY ──
    summary_df = pd.DataFrame([
        {'Sheet': '1 All Soluble Coffee',     'Rows': len(df_sheet1),        'Notes': 'Correct HSN + wrong HSN rescued rows'},
        {'Sheet': '2 Chicory Explicit Ratio', 'Rows': len(df_chic_explicit), 'Notes': 'Ratio or chicory word in product description'},
        {'Sheet': '3 Chicory Known Brand',    'Rows': len(df_chic_known),    'Notes': 'Matched to brand reference table'},
        {'Sheet': '4 Chicory Assumed',        'Rows': len(df_chic_assumed),  'Notes': 'Chicory signal but no confirmed ratio'},
        {'Sheet': '5 Chicory Only Exports',   'Rows': len(df_chicory_only),  'Notes': 'HSN 21013010 — pure roasted chicory exports'},
        {'Sheet': '6 Excluded Items Review',  'Rows': len(df_excluded_review), 'Notes': 'Non-soluble products removed'},
    ])

    return {
        '1 All Soluble Coffee':     clean_export(df_sheet1),
        '2 Chicory Explicit Ratio': clean_export(df_chic_explicit),
        '3 Chicory Known Brand':    clean_export(df_chic_known),
        '4 Chicory Assumed':        clean_export(df_chic_assumed),
        '5 Chicory Only Exports':   clean_export(df_chicory_only),
        '6 Excluded Items Review':  clean_export(df_excluded_review),
        'Summary':                  summary_df,
    }

# ================================================================
# EXCEL WRITER
# ================================================================
def write_excel(sheets_dict):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        for sheet_name, df_out in sheets_dict.items():
            df_out.to_excel(writer, sheet_name=sheet_name, index=False)

    buf.seek(0)
    wb = load_workbook(buf)
    for sheet_name, colours in SHEET_COLOURS.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            format_sheet(ws, *colours)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()

# ================================================================
# UI (UNCHANGED STRUCTURE)
# ================================================================
st.markdown('<div class="pipeline">', unsafe_allow_html=True)

st.subheader("Stage 1 — Upload Exclusion List")
excl = st.file_uploader("Exclusion List", type=["xlsx"])

st.subheader("Stage 2 — Upload Raw Data")
raws = st.file_uploader("Raw CYBEX Files", type=["xlsx"], accept_multiple_files=True)

if st.button("Run"):
    if not excl:
        st.error("Upload exclusion list")
    elif not raws:
        st.error("Upload files")
    else:
        excl_df = pd.read_excel(excl)
        # Normalise exclusion list columns
        excl_df.columns = excl_df.columns.str.strip().str.upper()
        if 'KEYWORD' not in excl_df.columns:
            st.error("Exclusion list must have a KEYWORD column")
        else:
            for f in raws:
                with st.spinner(f"Processing {f.name}..."):
                    result = process_file(f, excl_df)

                if result is None:
                    continue

                # Show summary
                st.success(f"✓ {f.name}")
                st.dataframe(result['Summary'], use_container_width=True)

                # Write formatted Excel
                excel_bytes = write_excel(result)
                out_name = f"CLEANED_{f.name}"

                st.download_button(
                    label=f"⬇ Download {out_name}",
                    data=excel_bytes,
                    file_name=out_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

st.markdown('</div>', unsafe_allow_html=True)
